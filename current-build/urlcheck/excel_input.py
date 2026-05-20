"""Excel input layer for the URL checker.

Responsibilities:
- Load a workbook (read-only) and enumerate sheets.
- Heuristically score columns for "URL-likeness" so the GUI / CLI can
  auto-pick a sensible column when the user doesn't specify one.
- Extract URLs from a chosen sheet+column, including hyperlink targets
  (which may differ from the displayed text).
- Record every cell where each URL appears (sheet name + cell coordinate).
- Deduplicate so the engine only checks each unique URL once.
- After Phase 1's engine returns results, map them back to all cell
  occurrences so callers (Phase 3 reporter) can list every location.

Design notes:
- We deliberately use openpyxl's read_only mode for memory efficiency on
  large workbooks (40k–100k rows is the target). read_only mode forfeits
  random-cell access and hyperlink visibility, so we have a "rich" pass
  (load_workbook with read_only=False) gated by `read_hyperlinks=True`.
  The default is hyperlink-aware because that's the more useful behavior
  for this app, with read_only as a fallback for very large files.
- Column references accept letters ("B"), 1-based indices ("2", 2), or
  header names matched in the configured header row.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Union
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from .models import UrlCheckResult
from .url_normalize import (
    NormalizationOptions,
    extract_urls_from_text,
    normalize_url,
)


# ---------------------------------------------------------------------------
# Typed exceptions
# ---------------------------------------------------------------------------
#
# Callers (the GUI worker, the CLI, tests) need to react differently to
# distinct failure modes — "file not found" wants a different message
# than "header X not in this sheet". Earlier versions of this module
# raised bare `ValueError` and `FileNotFoundError`, which forced callers
# to do brittle substring matching ("'not in workbook' in str(e)") to
# tell error categories apart. That broke quietly whenever this module's
# error wording was rephrased.
#
# Each exception below carries the relevant CONTEXT as attributes (the
# bad sheet name, the available sheets, the column reference, etc.) so
# callers can format their own messages without parsing the str(e).
#
# Backwards-compatible aliases:
# - `ExcelInputError` subclasses `ValueError` so old `except ValueError`
#   blocks still catch every category except file-missing.
# - `ExcelFileNotFound` subclasses BOTH `ExcelInputError` and the stdlib
#   `FileNotFoundError`, so existing `except FileNotFoundError` blocks
#   in callers continue to work unchanged.

class ExcelInputError(ValueError):
    """Base class for every error raised by `excel_input`.

    Inherits from `ValueError` for backwards compatibility — callers that
    catch `ValueError` to handle "any Excel-stage failure" keep working.
    New callers should catch this (or a specific subclass) instead.
    """


class ExcelFileNotFound(ExcelInputError, FileNotFoundError):
    """The Excel file path doesn't exist or isn't a regular file.

    Multiple-inherits from `FileNotFoundError` so legacy
    `except FileNotFoundError` paths still catch this. New code should
    prefer the typed form.
    """
    def __init__(self, path: Path):
        self.path = path
        super().__init__(f"Excel file not found: {path}")


class InvalidExcelFile(ExcelInputError):
    """The file exists but openpyxl rejected it (not a .xlsx).

    Common cause: an older .xls file that was renamed to .xlsx without
    actually converting it. Attribute `path` is the offending file;
    `cause_message` is whatever openpyxl's underlying exception said,
    in case the GUI wants to surface a "details" link.
    """
    def __init__(self, path: Path, cause_message: str = ""):
        self.path = path
        self.cause_message = cause_message
        super().__init__(f"Not a valid .xlsx file: {path}")


class EmptyWorkbook(ExcelInputError):
    """The workbook opened cleanly but contains zero sheets.

    Rare but possible (e.g. a workbook with every sheet deleted before
    saving). Raised by callers — `open_workbook` itself doesn't check
    this since openpyxl will load the file regardless.
    """
    def __init__(self, path: Path):
        self.path = path
        super().__init__(f"Workbook has no sheets: {path}")


class SheetNotFound(ExcelInputError):
    """The requested sheet name doesn't exist in the workbook.

    `sheet` is the missing name; `available` is the list of sheets that
    DO exist (so callers can suggest "did you mean X?" or list options).
    """
    def __init__(self, sheet: str, available: list[str]):
        self.sheet = sheet
        self.available = list(available)
        super().__init__(
            f"Sheet {sheet!r} not in workbook. "
            f"Available: {', '.join(self.available)}"
        )


class ColumnNotFound(ExcelInputError):
    """Base class for column-resolution failures."""


class HeaderNotFound(ColumnNotFound):
    """A header-name lookup failed.

    `header` is the name that was looked up; `sheet` and `header_row` are
    where we looked.
    """
    def __init__(self, header: str, sheet: str, header_row: int):
        self.header = header
        self.sheet = sheet
        self.header_row = header_row
        super().__init__(
            f"Header {header!r} not found in row {header_row} "
            f"of sheet {sheet!r}."
        )


class InvalidColumnReference(ColumnNotFound):
    """The column reference couldn't be parsed at all.

    Covers: negative or zero index, empty string, letters that aren't a
    valid Excel column (e.g. "@@"), and the case where a non-letter
    string is passed without a header row to look it up against.
    `column_ref` is the value the caller supplied.
    """
    def __init__(self, column_ref, reason: str):
        self.column_ref = column_ref
        self.reason = reason
        super().__init__(reason)


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class Occurrence:
    """One occurrence of a URL in the input workbook."""
    sheet: str
    cell: str  # A1-style coordinate, e.g. "B12"
    row: int
    col: int  # 1-based
    value: Optional[str] = None  # original cell display text, for debug
    hyperlink: Optional[str] = None  # hyperlink target if cell had one

    @property
    def location(self) -> str:
        """Stable string format used in reports: "Sheet!B12"."""
        # If the sheet name has spaces / punctuation, Excel formula style would
        # quote it ('Sheet 1'!B12). We keep the simpler form for human display
        # since the report is for colleagues, not formula evaluation.
        return f"{self.sheet}!{self.cell}"


@dataclass(frozen=True)
class ExcelSelection:
    """A resolved choice of which sheet+column to read.

    `column` is 1-based for normal column-targeted extractions. For
    whole-sheet scans (see `extract_urls_from_sheet`), `column` is set
    to the sentinel `WHOLE_SHEET_COLUMN` (0) and `column_letter` returns
    the marker `"*"` rather than raising. Callers that need a
    user-facing label should prefer `format_selection_label()`.
    """
    sheet: str
    column: int  # 1-based, or WHOLE_SHEET_COLUMN (0) for whole-sheet scans
    header: Optional[str] = None  # value found in the header row, if any

    @property
    def column_letter(self) -> str:
        # `get_column_letter(0)` raises — guard the sentinel here so
        # legacy callers that read `column_letter` without knowing about
        # the whole-sheet mode get a sensible placeholder back instead
        # of an exception.
        if self.column <= 0:
            return "*"
        return get_column_letter(self.column)


@dataclass
class ExtractionSummary:
    """Aggregate counts after extraction."""
    total_cells_scanned: int = 0
    total_occurrences: int = 0
    unique_urls: int = 0
    rejected_cells: int = 0  # cells with content that didn't yield a URL


@dataclass
class ColumnCandidate:
    """A column scored as a possible URL column."""
    column: int  # 1-based
    column_letter: str
    header: Optional[str]
    score: float  # 0..1
    sample_count: int  # how many sampled cells were URL-like


@dataclass
class ExtractionResult:
    """Full result of extracting URLs from a sheet+column."""
    selection: ExcelSelection
    unique_urls: list[str] = field(default_factory=list)
    occurrences_map: dict[str, list[Occurrence]] = field(default_factory=dict)
    summary: ExtractionSummary = field(default_factory=ExtractionSummary)


@dataclass
class EnrichedResult:
    """Pairs an engine UrlCheckResult with all the cell locations it covers."""
    result: UrlCheckResult
    occurrences: list[Occurrence]

    @property
    def locations(self) -> list[str]:
        return [o.location for o in self.occurrences]


# ---------------------------------------------------------------------------
# Workbook loading
# ---------------------------------------------------------------------------

def open_workbook(path: Union[str, Path], read_only: bool = False):
    """Open a workbook for reading.

    By default we use read_only=False so cell.hyperlink is populated. For
    very large workbooks where memory matters, callers can pass read_only=True
    at the cost of losing hyperlink visibility (we'll only see displayed text).

    Raises:
        ExcelFileNotFound: file does not exist.
        InvalidExcelFile: file exists but is not a valid .xlsx (this
            includes plain-text files renamed to .xlsx, corrupted
            archives, and older .xls binaries that were never converted).
    """
    p = Path(path)
    if not p.exists():
        raise ExcelFileNotFound(p)
    try:
        # data_only=True so cells with formulas yield their cached value,
        # not the formula text — important if URLs come from HYPERLINK() or
        # CONCAT() formulas.
        return load_workbook(filename=str(p), read_only=read_only, data_only=True)
    except InvalidFileException as e:
        raise InvalidExcelFile(p, cause_message=str(e)) from e
    except BadZipFile as e:
        # openpyxl validates the zip archive before parsing. A plain
        # text file or an old .xls binary trips BadZipFile, not
        # InvalidFileException — translate it the same way so the
        # caller sees one consistent error class regardless of how
        # the file is malformed.
        raise InvalidExcelFile(p, cause_message=str(e)) from e


def list_sheets(workbook) -> list[str]:
    """Return the workbook's sheet names in their natural order."""
    return list(workbook.sheetnames)


# ---------------------------------------------------------------------------
# Column reference resolution
# ---------------------------------------------------------------------------

ColumnRef = Union[int, str]


def resolve_column(
    workbook,
    sheet_name: str,
    column_ref: ColumnRef,
    header_row: Optional[int] = 1,
) -> ExcelSelection:
    """Resolve a user-provided column reference into an ExcelSelection.

    Accepts:
    - int (1-based column index), e.g. 2
    - str digits, e.g. "2"
    - str column letters, e.g. "B" or "AA"
    - str header name to look up in the header row, e.g. "URL"

    Raises:
        SheetNotFound: if the sheet_name doesn't exist.
        InvalidColumnReference: if the column reference can't be parsed
            (bad integer, empty string, bad letters, or non-letter text
            without a header row to look it up in).
        HeaderNotFound: if the reference is a header name but no row in
            the header row matches it.
    """
    ws = _get_sheet(workbook, sheet_name)

    # Case 1: int -> direct index.
    if isinstance(column_ref, int):
        if column_ref < 1:
            raise InvalidColumnReference(
                column_ref,
                f"Column index must be >= 1, got {column_ref}",
            )
        return ExcelSelection(
            sheet=sheet_name,
            column=column_ref,
            header=_read_header(ws, column_ref, header_row),
        )

    # Case 2: string. Try (in order): digits -> letters -> header name.
    s = str(column_ref).strip()
    if not s:
        raise InvalidColumnReference(column_ref, "Empty column reference")

    if s.isdigit():
        idx = int(s)
        if idx < 1:
            raise InvalidColumnReference(
                column_ref,
                f"Column index must be >= 1, got {idx}",
            )
        return ExcelSelection(
            sheet=sheet_name,
            column=idx,
            header=_read_header(ws, idx, header_row),
        )

    # Pure letters (A-Z, AA-ZZ, ...) — but a header row may have a single-letter
    # column name like "URL" which would otherwise be misread as column 14676.
    # Strategy: if header_row is set, try header-name lookup FIRST. If found,
    # use it. Otherwise fall back to column-letter interpretation.
    if re.fullmatch(r"[A-Za-z]+", s):
        if header_row and header_row >= 1:
            idx = _find_header_column(ws, s, header_row)
            if idx is not None:
                return ExcelSelection(sheet=sheet_name, column=idx, header=s)
        try:
            idx = column_index_from_string(s.upper())
        except ValueError as e:
            raise InvalidColumnReference(
                column_ref, f"Invalid column letter: {s!r}"
            ) from e
        return ExcelSelection(
            sheet=sheet_name,
            column=idx,
            header=_read_header(ws, idx, header_row),
        )

    # Otherwise treat as header name (the string contains digits, spaces, or
    # punctuation, so it's definitely not a column letter).
    if header_row is None or header_row < 1:
        raise InvalidColumnReference(
            column_ref,
            f"Column {s!r} doesn't look like an index or letter, and "
            "no header row is set to look up the header by name.",
        )
    idx = _find_header_column(ws, s, header_row)
    if idx is None:
        raise HeaderNotFound(s, sheet_name, header_row)
    return ExcelSelection(sheet=sheet_name, column=idx, header=s)


def _get_sheet(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        raise SheetNotFound(sheet_name, list(workbook.sheetnames))
    return workbook[sheet_name]


def _read_header(ws, column_index: int, header_row: Optional[int]) -> Optional[str]:
    if not header_row or header_row < 1:
        return None
    cell = ws.cell(row=header_row, column=column_index)
    val = cell.value
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def _find_header_column(ws, header_name: str, header_row: int) -> Optional[int]:
    target = header_name.strip().lower()
    # max_column may be None in read_only mode without iter; iter_rows is safe.
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            if str(cell.value).strip().lower() == target:
                return cell.column
        return None  # only one row to scan
    return None


# ---------------------------------------------------------------------------
# Column detection (URL-likeness scoring)
# ---------------------------------------------------------------------------

def detect_url_columns(
    workbook,
    sheet_name: str,
    scan_rows: int = 50,
    header_row: Optional[int] = 1,
) -> list[ColumnCandidate]:
    """Score every column by how URL-like its values look.

    Returns candidates sorted by score descending. Score is the fraction of
    sampled cells whose value looks like a URL (or is a hyperlink). The
    result list always includes every column that had at least one URL-like
    cell, so the caller can pick the top one or display them all.

    "URL-like" means any of:
    - cell.hyperlink.target is a usable http(s) URL
    - cell value, when run through `extract_urls_from_text`, yields >= 1 URL
    """
    ws = _get_sheet(workbook, sheet_name)

    # Determine the row range to scan.
    start_row = (header_row + 1) if header_row else 1
    end_row = start_row + scan_rows - 1

    # Per-column counters: { col_index: [hits, total_nonempty] }
    counters: dict[int, list[int]] = {}

    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        values_only=False,
    ):
        for cell in row:
            value = cell.value
            hyperlink_target = _hyperlink_target(cell)
            if value is None and hyperlink_target is None:
                continue
            col = cell.column
            slot = counters.setdefault(col, [0, 0])
            slot[1] += 1  # nonempty

            if hyperlink_target and normalize_url(hyperlink_target):
                slot[0] += 1
                continue
            if value is not None and extract_urls_from_text(str(value)):
                slot[0] += 1

    candidates: list[ColumnCandidate] = []
    for col, (hits, total) in sorted(counters.items()):
        if hits == 0:
            continue
        score = hits / total if total else 0.0
        candidates.append(
            ColumnCandidate(
                column=col,
                column_letter=get_column_letter(col),
                header=_read_header(ws, col, header_row),
                score=score,
                sample_count=hits,
            )
        )

    # Sort: highest score first, then by sample_count to break ties, then by column.
    candidates.sort(key=lambda c: (-c.score, -c.sample_count, c.column))
    return candidates


def auto_pick_column(
    workbook,
    sheet_name: str,
    scan_rows: int = 50,
    header_row: Optional[int] = 1,
) -> Optional[ExcelSelection]:
    """Convenience: pick the best URL-y column for a sheet, or None if none."""
    candidates = detect_url_columns(workbook, sheet_name, scan_rows, header_row)
    if not candidates:
        return None
    top = candidates[0]
    return ExcelSelection(sheet=sheet_name, column=top.column, header=top.header)


def auto_pick_sheet_and_column(
    workbook,
    scan_rows: int = 50,
    header_row: Optional[int] = 1,
) -> Optional[ExcelSelection]:
    """Pick the best (sheet, column) pair across all sheets.

    Returns the candidate with the highest score; ties broken by sample_count.
    Useful when the user provided neither --sheet nor --column.
    """
    best: Optional[ExcelSelection] = None
    best_score: tuple[float, int] = (-1.0, -1)
    for sheet in workbook.sheetnames:
        cands = detect_url_columns(workbook, sheet, scan_rows, header_row)
        if not cands:
            continue
        top = cands[0]
        key = (top.score, top.sample_count)
        if key > best_score:
            best_score = key
            best = ExcelSelection(sheet=sheet, column=top.column, header=top.header)
    return best


# ---------------------------------------------------------------------------
# Hyperlink helper
# ---------------------------------------------------------------------------

def _hyperlink_target(cell) -> Optional[str]:
    """Return cell.hyperlink.target if present and string-like, else None.

    openpyxl exposes hyperlinks as a Hyperlink object with `.target`.
    Internal-only links (sheet anchors) are skipped.
    """
    hl = getattr(cell, "hyperlink", None)
    if hl is None:
        return None
    target = getattr(hl, "target", None)
    if not target:
        # Some hyperlinks are internal (sheet jumps) — those have a .location
        # but no .target. We don't want those.
        return None
    return str(target)


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

def extract_urls(
    workbook,
    selection: ExcelSelection,
    header_row: Optional[int] = 1,
    options: Optional[NormalizationOptions] = None,
) -> ExtractionResult:
    """Extract URLs from `selection.sheet` / `selection.column`.

    For each cell in the column (skipping the header row if set):
    1. If the cell carries a hyperlink, extract its target.
    2. Run the cell's display text through `extract_urls_from_text`.
    3. Combine + deduplicate within-cell, then record an Occurrence per URL.

    Hyperlink targets are normalized through the same pipeline as text URLs,
    so a `mailto:` hyperlink (for example) will still be rejected.
    """
    ws = _get_sheet(workbook, selection.sheet)
    options = options or NormalizationOptions()

    summary = ExtractionSummary()
    occurrences_map: dict[str, list[Occurrence]] = {}
    unique_order: list[str] = []  # preserve first-seen order for reproducibility

    start_row = (header_row + 1) if header_row else 1

    # We iterate the whole column. iter_rows with min/max_col=column is
    # cheaper than iter_cols for read-only workbooks.
    for row in ws.iter_rows(
        min_row=start_row,
        min_col=selection.column,
        max_col=selection.column,
        values_only=False,
    ):
        cell = row[0]
        value = cell.value
        hyperlink_target = _hyperlink_target(cell)

        if value is None and hyperlink_target is None:
            continue

        summary.total_cells_scanned += 1

        urls_in_cell: list[str] = []

        # 1) Hyperlink target first (often the "real" URL even when display
        #    text is something like "click here").
        if hyperlink_target is not None:
            normalized = normalize_url(hyperlink_target, options)
            if normalized:
                urls_in_cell.append(normalized)

        # 2) Display text (may carry additional URLs or the only URL if no hyperlink).
        if value is not None:
            urls_in_cell.extend(extract_urls_from_text(str(value), options))

        # Dedupe within the cell while preserving order.
        seen_in_cell: set[str] = set()
        cell_unique: list[str] = []
        for u in urls_in_cell:
            if u not in seen_in_cell:
                seen_in_cell.add(u)
                cell_unique.append(u)

        if not cell_unique:
            summary.rejected_cells += 1
            continue

        original_value = None if value is None else str(value)
        for url in cell_unique:
            occ = Occurrence(
                sheet=selection.sheet,
                cell=cell.coordinate,
                row=cell.row,
                col=cell.column,
                value=original_value,
                hyperlink=hyperlink_target,
            )
            if url not in occurrences_map:
                occurrences_map[url] = []
                unique_order.append(url)
            occurrences_map[url].append(occ)
            summary.total_occurrences += 1

    summary.unique_urls = len(unique_order)
    return ExtractionResult(
        selection=selection,
        unique_urls=unique_order,
        occurrences_map=occurrences_map,
        summary=summary,
    )


# ---------------------------------------------------------------------------
# Whole-sheet extraction (no column picking required)
# ---------------------------------------------------------------------------

# Sentinel used by `ExcelSelection.column` when we did a whole-sheet scan.
# 0 isn't a valid 1-based column index, so it can never collide with a real
# column. We expose `WHOLE_SHEET_COLUMN` as a public name and override
# `column_letter` lookups to handle it gracefully.
WHOLE_SHEET_COLUMN: int = 0


def _whole_sheet_selection(sheet_name: str) -> ExcelSelection:
    """Build the ExcelSelection placeholder used for whole-sheet scans.

    Real column-based extractions have a meaningful column index and
    optional header. A whole-sheet scan doesn't, so we use sentinel
    values that downstream code can recognise:
      - column = WHOLE_SHEET_COLUMN (0)
      - header = None
    Callers that print `selection.column_letter` need to special-case
    the sentinel — see `format_selection_label` below.
    """
    return ExcelSelection(
        sheet=sheet_name,
        column=WHOLE_SHEET_COLUMN,
        header=None,
    )


def format_selection_label(selection: ExcelSelection) -> str:
    """Human-friendly label for a selection (handles the whole-sheet case).

    Used by the GUI and the report layer so they don't each have to
    decide what to print when `column == WHOLE_SHEET_COLUMN`.
    """
    if selection.column == WHOLE_SHEET_COLUMN:
        return f"{selection.sheet} (all columns)"
    return f"{selection.sheet}!{selection.column_letter}"


def extract_urls_from_sheet(
    workbook,
    sheet_name: str,
    header_row: Optional[int] = 1,
    options: Optional[NormalizationOptions] = None,
) -> ExtractionResult:
    """Extract URLs from EVERY cell in `sheet_name`, regardless of column.

    This is the "scan everything" path used when the user (or the GUI's
    default mode) doesn't want to pick a specific URL column. We walk
    every row and every column, applying exactly the same per-cell
    rules as `extract_urls`:

    1. If the cell carries a hyperlink, the target is extracted and
       normalized.
    2. The cell's display text is run through `extract_urls_from_text`
       so multi-URL cells (whitespace / comma / semicolon separated)
       still yield every URL.
    3. URLs are deduplicated within the cell, then recorded as an
       Occurrence per (URL, cell) pair.

    The header row, if set, is skipped on EVERY column — not just one —
    so a "URL" or "Link" header cell with no hyperlink doesn't pollute
    the unique-URL list. (If a header cell carries a hyperlink, that's
    almost certainly a real link the user wants checked, so we still
    record it; see the per-row guard below.)

    Returned `ExtractionResult.selection.column` is set to the sentinel
    `WHOLE_SHEET_COLUMN` (0) so callers can tell the two extraction
    modes apart. `column_letter` is not meaningful in this mode — use
    `format_selection_label(result.selection)` for display.

    Notes on cost:
    - openpyxl's `iter_rows()` is row-streaming and only materialises
      cells that exist in the sheet's used range, so a workbook with a
      lot of empty trailing columns / rows doesn't pay for them.
    - We still pre-resolve `header_row` to a comparison int so the
      per-row check is a single integer compare, not an attribute fetch
      per cell.
    """
    ws = _get_sheet(workbook, sheet_name)
    options = options or NormalizationOptions()

    summary = ExtractionSummary()
    occurrences_map: dict[str, list[Occurrence]] = {}
    unique_order: list[str] = []  # preserve first-seen order

    # Normalise header_row to either a positive int (skip exactly that row)
    # or None (skip nothing). Anything <1 means "no header" — Excel rows
    # are 1-based.
    skip_row = header_row if (header_row and header_row >= 1) else None

    for row in ws.iter_rows(values_only=False):
        for cell in row:
            # Skip header cells, but only if they have no hyperlink. A
            # hyperlinked header cell is almost certainly a real URL the
            # user added and wants checked.
            if skip_row is not None and cell.row == skip_row:
                if _hyperlink_target(cell) is None:
                    continue

            value = cell.value
            hyperlink_target = _hyperlink_target(cell)
            if value is None and hyperlink_target is None:
                continue

            summary.total_cells_scanned += 1

            urls_in_cell: list[str] = []

            # 1) Hyperlink target first (often the "real" URL even when
            #    display text is something like "click here").
            if hyperlink_target is not None:
                normalized = normalize_url(hyperlink_target, options)
                if normalized:
                    urls_in_cell.append(normalized)

            # 2) Display text (may carry additional URLs or be the only
            #    URL source if there's no hyperlink).
            if value is not None:
                urls_in_cell.extend(extract_urls_from_text(str(value), options))

            # Dedupe within the cell while preserving order.
            seen_in_cell: set[str] = set()
            cell_unique: list[str] = []
            for u in urls_in_cell:
                if u not in seen_in_cell:
                    seen_in_cell.add(u)
                    cell_unique.append(u)

            if not cell_unique:
                summary.rejected_cells += 1
                continue

            original_value = None if value is None else str(value)
            for url in cell_unique:
                occ = Occurrence(
                    sheet=sheet_name,
                    cell=cell.coordinate,
                    row=cell.row,
                    col=cell.column,
                    value=original_value,
                    hyperlink=hyperlink_target,
                )
                if url not in occurrences_map:
                    occurrences_map[url] = []
                    unique_order.append(url)
                occurrences_map[url].append(occ)
                summary.total_occurrences += 1

    summary.unique_urls = len(unique_order)
    return ExtractionResult(
        selection=_whole_sheet_selection(sheet_name),
        unique_urls=unique_order,
        occurrences_map=occurrences_map,
        summary=summary,
    )


# ---------------------------------------------------------------------------
# Mapping engine results back to cell locations
# ---------------------------------------------------------------------------

def enrich_results_with_occurrences(
    results: list[UrlCheckResult],
    occurrences_map: dict[str, list[Occurrence]],
) -> list[EnrichedResult]:
    """For each engine result, attach the list of Occurrences for that URL.

    Engine results are keyed by `original_url`, which is the same normalized
    string we used as the occurrences_map key. URLs in the engine results
    that aren't in the map (shouldn't happen, but be defensive) get an empty
    occurrences list rather than raising.
    """
    enriched: list[EnrichedResult] = []
    for r in results:
        occs = occurrences_map.get(r.original_url, [])
        enriched.append(EnrichedResult(result=r, occurrences=list(occs)))
    return enriched
