"""Excel output layer for the URL checker.

Generates a focused, two-sheet "issues" report from Phase 2's enriched
results. Only problematic URLs are included — OK results are omitted by
design. The report is intended for non-technical reviewers, so the file
is sorted by domain (so domain-level blocking patterns jump out), header
rows are styled and frozen, and column widths are auto-fit.

Public API:
    build_output_rows(enriched) -> (broken_rows, blocked_rows)
    sort_rows(rows)             -> rows sorted by (domain, url)
    write_excel_report(broken_rows, blocked_rows, file_path)
    auto_adjust_columns(ws)
    style_header(ws)
    default_output_filename()   -> "url_issues_YYYY-MM-DD.xlsx"

Design choices (documented):
- Filename uses local date (per spec). Times-of-day are deliberately
  excluded so a same-day re-run produces the same path; users can pass
  --output to override if they want timestamped names.
- "Cell Location(s)" is a comma-separated string. Excel doesn't support
  arrays in cells, and a single sortable string is what colleagues
  actually want to scan.
- Domain extraction uses urllib.parse.urlparse(...).netloc lowercased,
  with port stripped, so "example.com:443" and "example.com" sort
  together.
- We use a professional, fixed font (Arial) for consistency.
- Banded row fills are intentionally NOT applied — they look noisy on
  reports with mixed text lengths and add no information. The frozen
  bold header is enough.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable, Optional, Sequence
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import Classification, UrlCheckResult


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SHEET_BROKEN = "Broken URLs"
SHEET_BLOCKED = "Possibly Blocked"
SHEET_ALL_CLEAR = "All Clear"

# Columns in EXACT order (per spec).
BROKEN_COLUMNS: tuple[str, ...] = (
    "URL",
    "Cell Location(s)",
    "HTTP Code",
    "Error Detail",
    "Final URL",
    "Response Time (ms)",
    "Checked At (UTC)",
)

BLOCKED_COLUMNS: tuple[str, ...] = BROKEN_COLUMNS + ("Likely Reason",)

# Column-name → row-key mapping. Centralized so the field key changes
# in one place if a column header text is ever rephrased.
_KEY_BY_COLUMN: dict[str, str] = {
    "URL": "url",
    "Cell Location(s)": "cell_locations",
    "HTTP Code": "http_status",
    "Error Detail": "error_detail",
    "Final URL": "final_url",
    "Response Time (ms)": "response_time_ms",
    "Checked At (UTC)": "checked_at_utc",
    "Likely Reason": "likely_reason",
}

# Columns that should be center-aligned (codes/times). Everything else is
# left-aligned — appropriate for free-text URLs and locations.
_CENTERED_COLUMNS: frozenset[str] = frozenset({
    "HTTP Code", "Response Time (ms)", "Checked At (UTC)",
})

# Styling.
_FONT_NAME = "Arial"
_HEADER_FONT = Font(name=_FONT_NAME, bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill("solid", start_color="FF305496", end_color="FF305496")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center")
_BODY_FONT = Font(name=_FONT_NAME)
_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=False)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
_ALL_CLEAR_FONT = Font(name=_FONT_NAME, bold=True, size=12)

# Auto-width tuning.
_AUTOWIDTH_PADDING = 2
_AUTOWIDTH_MIN = 10
_AUTOWIDTH_MAX = 80  # cap so an enormous URL doesn't blow out the column


# ---------------------------------------------------------------------------
# Filename helper
# ---------------------------------------------------------------------------

def default_output_filename(today: Optional[date] = None) -> str:
    """Return the spec-mandated filename: url_issues_YYYY-MM-DD.xlsx (local date)."""
    d = today or date.today()
    return f"url_issues_{d.isoformat()}.xlsx"


# ---------------------------------------------------------------------------
# Domain extraction (reused from engine logic conceptually, kept independent
# here so excel_output has no dependency on engine)
# ---------------------------------------------------------------------------

def _domain_for_sort(url: str) -> str:
    """Lowercased hostname (without port) for sorting purposes.

    Falls back to the raw URL string if parsing yields nothing usable, so
    pathological inputs still sort deterministically rather than crashing.
    """
    try:
        host = urlparse(url).hostname or ""
    except Exception:
        host = ""
    return host.lower() or url.lower()


# ---------------------------------------------------------------------------
# Row building
# ---------------------------------------------------------------------------

def _format_cell_locations(locations: Sequence[str]) -> str:
    """Join cell-location strings with ", ". Order is preserved from input."""
    return ", ".join(locations)


def _row_for_result(
    result: UrlCheckResult,
    locations: Sequence[str],
    *,
    include_likely_reason: bool,
) -> dict:
    """Build a row dict for one enriched result.

    Keys match `_KEY_BY_COLUMN` values so `_write_rows` can pull the right
    cell value per column.
    """
    row: dict = {
        "url": result.original_url,
        "cell_locations": _format_cell_locations(locations),
        "http_status": result.http_status if result.http_status is not None else "",
        "error_detail": result.error_detail or "",
        "final_url": result.final_url or "",
        "response_time_ms": result.response_time_ms if result.response_time_ms is not None else "",
        "checked_at_utc": result.checked_at_utc or "",
        # Sort keys (not written to the sheet).
        "_domain": _domain_for_sort(result.original_url),
        "_url_lower": result.original_url.lower(),
    }
    if include_likely_reason:
        row["likely_reason"] = result.likely_reason or ""
    return row


def build_output_rows(
    enriched: Iterable,
) -> tuple[list[dict], list[dict]]:
    """Split enriched results into (broken_rows, blocked_rows).

    `enriched` is iterable of objects with `.result` (UrlCheckResult) and
    `.locations` (list[str]) attributes — i.e. EnrichedResult from
    excel_input.py. Duck-typing rather than a hard import so this module
    can also be fed plain dicts (handy for tests / Phase 4 GUI).

    Filter: OK results are dropped. Only BROKEN goes to sheet 1 and only
    POSSIBLY_BLOCKED goes to sheet 2.
    """
    broken_rows: list[dict] = []
    blocked_rows: list[dict] = []

    for item in enriched:
        result, locations = _coerce_item(item)
        cls = result.classification
        if cls == Classification.BROKEN:
            broken_rows.append(
                _row_for_result(result, locations, include_likely_reason=False)
            )
        elif cls == Classification.POSSIBLY_BLOCKED:
            blocked_rows.append(
                _row_for_result(result, locations, include_likely_reason=True)
            )
        # OK and any unknown classifications are dropped.

    return broken_rows, blocked_rows


def _coerce_item(item) -> tuple[UrlCheckResult, list[str]]:
    """Accept either an EnrichedResult-like object or a (result, locations) tuple."""
    if hasattr(item, "result") and hasattr(item, "locations"):
        return item.result, list(item.locations)
    if isinstance(item, tuple) and len(item) == 2:
        return item[0], list(item[1])
    raise TypeError(
        "Expected EnrichedResult or (UrlCheckResult, list[str]); "
        f"got {type(item).__name__}"
    )


# ---------------------------------------------------------------------------
# Sorting
# ---------------------------------------------------------------------------

def sort_rows(rows: list[dict]) -> list[dict]:
    """Sort rows by (domain asc, url asc).

    Stable: equal-domain rows stay in their original relative order if
    URLs also tie — though in practice URL ties are unlikely after
    deduplication.
    """
    return sorted(rows, key=lambda r: (r.get("_domain", ""), r.get("_url_lower", "")))


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def style_header(ws: Worksheet) -> None:
    """Apply bold, white-on-blue, left-aligned styling to row 1.

    Also sets a slightly taller header row so the bold text breathes.
    """
    if ws.max_row < 1:
        return
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 20


def auto_adjust_columns(ws: Worksheet) -> None:
    """Set column widths based on the longest visible value in each column.

    Bounded by `_AUTOWIDTH_MIN` and `_AUTOWIDTH_MAX` so a giant URL doesn't
    push the column off-screen and an empty column doesn't collapse.
    """
    if ws.max_column == 0:
        return
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            value = row[0]
            if value is None:
                continue
            length = len(str(value))
            if length > max_len:
                max_len = length
        width = min(_AUTOWIDTH_MAX, max(_AUTOWIDTH_MIN, max_len + _AUTOWIDTH_PADDING))
        ws.column_dimensions[letter].width = width


def _apply_body_alignment(ws: Worksheet, columns: Sequence[str]) -> None:
    """Apply per-column body alignment (centered for codes/times, left for text).

    Skips the header row, which has its own alignment from `style_header`.
    """
    if ws.max_row < 2:
        return
    for col_idx, header in enumerate(columns, start=1):
        align = _CENTER_ALIGN if header in _CENTERED_COLUMNS else _LEFT_ALIGN
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            cell.font = _BODY_FONT
            cell.alignment = align


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def _write_sheet(
    ws: Worksheet,
    columns: Sequence[str],
    rows: Sequence[dict],
) -> None:
    """Write headers + data rows to a worksheet, then style + freeze + size."""
    # Header.
    for col_idx, header in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Data.
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, header in enumerate(columns, start=1):
            key = _KEY_BY_COLUMN[header]
            ws.cell(row=r_idx, column=c_idx, value=row.get(key, ""))

    # Styling.
    style_header(ws)
    _apply_body_alignment(ws, columns)
    ws.freeze_panes = "A2"
    auto_adjust_columns(ws)


def _write_all_clear_sheet(ws: Worksheet) -> None:
    """Single-sheet "no issues" report."""
    ws.cell(
        row=1, column=1,
        value="\u2705 No URL issues found. All links returned HTTP 200.",
    )
    ws["A1"].font = _ALL_CLEAR_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    # Generous width so the whole sentence is visible without squinting.
    ws.column_dimensions["A"].width = 80
    ws.row_dimensions[1].height = 22


def write_excel_report(
    broken_rows: Sequence[dict],
    blocked_rows: Sequence[dict],
    file_path: str | Path,
) -> Path:
    """Write the report file. Returns the resolved Path that was written.

    If both row lists are empty, an "All Clear" workbook is written instead
    of the two-sheet structure (per spec).
    """
    out_path = Path(file_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # The template-default first sheet — we'll rename / repurpose below.
    default_sheet = wb.active

    if not broken_rows and not blocked_rows:
        default_sheet.title = SHEET_ALL_CLEAR
        _write_all_clear_sheet(default_sheet)
        wb.save(out_path)
        return out_path

    # Two-sheet structure (in spec order: Broken, then Blocked).
    # Important: sort each list independently so each sheet is internally
    # ordered by (domain, url) — sorting both together would interleave.
    broken_sorted = sort_rows(list(broken_rows))
    blocked_sorted = sort_rows(list(blocked_rows))

    default_sheet.title = SHEET_BROKEN
    _write_sheet(default_sheet, BROKEN_COLUMNS, broken_sorted)

    blocked_ws = wb.create_sheet(SHEET_BLOCKED)
    _write_sheet(blocked_ws, BLOCKED_COLUMNS, blocked_sorted)

    wb.save(out_path)
    return out_path
