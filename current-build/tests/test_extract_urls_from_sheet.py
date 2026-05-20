"""Tests for `extract_urls_from_sheet` — the whole-sheet extraction path.

`extract_urls_from_sheet` is the alternative to `extract_urls` used when
the caller doesn't want to nominate a specific URL column. It must
walk every cell on the chosen sheet and apply the same per-cell URL
extraction rules as the column-targeted path.

These tests cover:
- URLs in any column are picked up
- URLs in multiple columns of the same row are all picked up
- The header row is skipped (unless a header cell carries a hyperlink)
- Hyperlinked cells whose display text is "click here" yield the
  hyperlink target, not the display text
- Multi-URL cells (whitespace / comma / semicolon separated) yield
  every URL
- Cells whose contents look nothing like a URL are rejected, not
  silently kept
- Within-cell duplicates collapse; across-cell occurrences are kept
- The returned ExcelSelection uses the WHOLE_SHEET_COLUMN sentinel and
  format_selection_label renders it cleanly

Run with:
    python -m unittest tests.test_extract_urls_from_sheet -v
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from urlcheck.excel_input import (
    WHOLE_SHEET_COLUMN,
    ExcelSelection,
    extract_urls_from_sheet,
    format_selection_label,
    open_workbook,
)


def _save(wb, path: Path) -> None:
    wb.save(path)


class FindsUrlsAcrossColumns(unittest.TestCase):
    """URLs scattered across different columns should all be found."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "input.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_urls_in_different_columns_are_all_extracted(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Row 1 is the header row — should be skipped.
        ws.append(["Title", "Source", "Notes"])
        # URL in col A, B, and C across different rows.
        ws.append(["https://aaa.example/page", "manual entry", "n/a"])
        ws.append(["plain text", "https://bbb.example/x", "plain text"])
        ws.append(["plain text", "plain text", "https://ccc.example/y"])
        _save(wb, self.path)

        loaded = open_workbook(self.path)
        result = extract_urls_from_sheet(loaded, "Sheet1", header_row=1)

        self.assertEqual(
            sorted(result.unique_urls),
            sorted([
                "https://aaa.example/page",
                "https://bbb.example/x",
                "https://ccc.example/y",
            ]),
        )
        self.assertEqual(result.summary.unique_urls, 3)
        # Three URL-bearing cells contributed one occurrence each.
        self.assertEqual(result.summary.total_occurrences, 3)

    def test_multiple_urls_in_same_row_across_columns(self):
        """A row with URLs in several columns should yield every URL."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Links"
        ws.append(["Vendor", "Site", "Docs"])
        ws.append([
            "Acme",
            "https://acme.example",
            "https://docs.acme.example/intro",
        ])
        _save(wb, self.path)

        loaded = open_workbook(self.path)
        result = extract_urls_from_sheet(loaded, "Links", header_row=1)

        self.assertEqual(
            sorted(result.unique_urls),
            ["https://acme.example", "https://docs.acme.example/intro"],
        )
        # Verify each URL is mapped back to the right cell coordinate.
        occs_site = result.occurrences_map["https://acme.example"]
        self.assertEqual(len(occs_site), 1)
        self.assertEqual(occs_site[0].cell, "B2")
        occs_docs = result.occurrences_map["https://docs.acme.example/intro"]
        self.assertEqual(occs_docs[0].cell, "C2")


class HeaderRowHandling(unittest.TestCase):
    """Verify the header row is skipped — but hyperlinked header cells
    are still honored.
    """

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "input.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_header_row_contents_are_not_extracted_when_skipped(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # The header cells happen to LOOK like URLs (a corner case);
        # they must still be skipped because header_row=1.
        ws.append(["https://header.example/should-not-appear", "Other"])
        ws.append(["https://body.example/a", "plain text"])
        _save(wb, self.path)

        loaded = open_workbook(self.path)
        result = extract_urls_from_sheet(loaded, "Sheet1", header_row=1)

        self.assertEqual(result.unique_urls, ["https://body.example/a"])
        # No "header.example" anywhere in the occurrences.
        self.assertNotIn(
            "https://header.example/should-not-appear",
            result.occurrences_map,
        )

    def test_no_header_row_means_every_row_is_extracted(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["https://row1.example/x"])
        ws.append(["https://row2.example/x"])
        _save(wb, self.path)

        loaded = open_workbook(self.path)
        # header_row=None or 0 = "no header row, scan everything".
        result = extract_urls_from_sheet(loaded, "Sheet1", header_row=None)

        self.assertEqual(
            sorted(result.unique_urls),
            ["https://row1.example/x", "https://row2.example/x"],
        )


class WithinCellMultiUrlAndDedup(unittest.TestCase):
    """A single cell containing several URLs should yield each one."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "input.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_comma_separated_urls_in_one_cell(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Links"])
        ws.append(["https://a.example, https://b.example; https://c.example"])
        _save(wb, self.path)

        loaded = open_workbook(self.path)
        result = extract_urls_from_sheet(loaded, "Sheet1", header_row=1)

        self.assertEqual(
            sorted(result.unique_urls),
            ["https://a.example", "https://b.example", "https://c.example"],
        )

    def test_duplicate_urls_across_cells_collapse_to_one_unique(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["URL"])
        ws.append(["https://dup.example"])
        ws.append(["https://dup.example"])
        ws.append(["https://other.example"])
        _save(wb, self.path)

        loaded = open_workbook(self.path)
        result = extract_urls_from_sheet(loaded, "Sheet1", header_row=1)

        self.assertEqual(result.summary.unique_urls, 2)
        self.assertEqual(result.summary.total_occurrences, 3)
        # Both occurrences of the duplicate URL are recorded.
        self.assertEqual(len(result.occurrences_map["https://dup.example"]), 2)


class HyperlinkAwareness(unittest.TestCase):
    """Cells with hyperlinks should yield the hyperlink target."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "input.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_hyperlink_target_preferred_when_display_is_friendly_text(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Source"])
        # Display text is "click here" — hyperlink carries the real URL.
        cell = ws.cell(row=2, column=1, value="click here")
        cell.hyperlink = "https://hyperlinked.example/page"
        _save(wb, self.path)

        loaded = open_workbook(self.path)
        result = extract_urls_from_sheet(loaded, "Sheet1", header_row=1)

        self.assertEqual(
            result.unique_urls, ["https://hyperlinked.example/page"]
        )

    def test_non_url_cells_are_rejected(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["URL"])
        ws.append(["just some plain text"])
        ws.append(["N/A"])
        ws.append([42])  # numeric cell, no URL
        _save(wb, self.path)

        loaded = open_workbook(self.path)
        result = extract_urls_from_sheet(loaded, "Sheet1", header_row=1)

        self.assertEqual(result.unique_urls, [])
        # The non-URL cells were counted but rejected.
        self.assertEqual(result.summary.unique_urls, 0)
        self.assertGreaterEqual(result.summary.rejected_cells, 3)


class SelectionLabel(unittest.TestCase):
    """The returned selection should use the whole-sheet sentinel and
    `format_selection_label` should render it cleanly.
    """

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "input.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_selection_column_is_whole_sheet_sentinel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "MySheet"
        ws.append(["URL"])
        ws.append(["https://example.com"])
        _save(wb, self.path)

        loaded = open_workbook(self.path)
        result = extract_urls_from_sheet(loaded, "MySheet", header_row=1)

        self.assertEqual(result.selection.column, WHOLE_SHEET_COLUMN)
        # column_letter must not raise — it should return the placeholder.
        self.assertEqual(result.selection.column_letter, "*")
        # And the human-friendly label.
        self.assertEqual(
            format_selection_label(result.selection),
            "MySheet (all columns)",
        )

    def test_format_selection_label_for_normal_column(self):
        sel = ExcelSelection(sheet="Tab", column=3, header="Link")
        self.assertEqual(format_selection_label(sel), "Tab!C")


if __name__ == "__main__":
    unittest.main()
