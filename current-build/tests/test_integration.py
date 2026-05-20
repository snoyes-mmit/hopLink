"""End-to-end integration test for the extract → check → write pipeline.

We build a synthetic input workbook in a tempdir, run Phase 2 extraction
on it, swap in canned `UrlCheckResult` objects for the HTTP layer (the
engine itself is exercised by the unit tests in `test_signatures.py` —
running it for real would require network access we don't have in CI),
then run Phase 2's enrichment step and Phase 3's writer, and finally
verify the output workbook is shaped correctly.

This is the test that would have caught the closing-paren bug from
priority item #1 if it had existed: the input workbook includes
Wikipedia-style URLs with embedded parens, and the output sheet's
"URL" column should preserve them.

Run with:
    python -m unittest tests.test_integration -v
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from urlcheck.excel_input import (
    enrich_results_with_occurrences,
    extract_urls,
    open_workbook,
    resolve_column,
)
from urlcheck.excel_output import (
    SHEET_ALL_CLEAR,
    SHEET_BLOCKED,
    SHEET_BROKEN,
    build_output_rows,
    write_excel_report,
)
from urlcheck.models import Classification, UrlCheckResult


def _result(url, classification, **kwargs):
    """Build a `UrlCheckResult` for the canned engine."""
    defaults = dict(
        original_url=url,
        domain=url.split("//", 1)[-1].split("/", 1)[0],
        classification=classification,
        final_url=url,
        http_status=None,
        error_detail=None,
        likely_reason=None,
        response_time_ms=100,
        checked_at_utc="2026-05-18T12:00:00Z",
    )
    defaults.update(kwargs)
    return UrlCheckResult(**defaults)


def _build_input_workbook(path: Path, rows: list[list[object]]) -> None:
    """Write a one-sheet workbook with [URL] header + given rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["URL"])
    for row in rows:
        ws.append(row)
    wb.save(path)


# ---------------------------------------------------------------------------
# End-to-end pipeline test
# ---------------------------------------------------------------------------

class FullPipelineHappyPath(unittest.TestCase):
    """Run extract → (canned results) → enrich → write and verify output."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self.tmp.name)
        self.input_path = self.tmp_path / "input.xlsx"
        self.output_path = self.tmp_path / "output.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_pipeline_produces_correct_report(self):
        # Input: a mix of OK, broken, blocked, and edge-case URLs.
        # The Wikipedia URL with parens is the regression-guard from item #1.
        _build_input_workbook(self.input_path, [
            ["https://ok-site.com/page"],                                # OK
            ["https://broken-site.com/missing"],                          # BROKEN 404
            ["https://throttled-site.com/api"],                           # BLOCKED 429
            ["https://en.wikipedia.org/wiki/Python_(programming_language)"],  # OK + parens
            ["https://dead-domain.example/x"],                            # BROKEN no-conn
            ["N/A"],                                                       # rejected
            [""],                                                          # empty
            ["https://broken-site.com/missing"],                           # duplicate
        ])

        # --- Phase 2: extract from the workbook ---
        wb = open_workbook(self.input_path)
        selection = resolve_column(wb, "Sheet1", "URL", header_row=1)
        extraction = extract_urls(wb, selection, header_row=1)

        # Unique URLs (rejects filtered, duplicates deduplicated):
        self.assertEqual(extraction.summary.unique_urls, 5)
        # Total occurrences (duplicate counted): 5 unique + 1 dup = 6.
        self.assertEqual(extraction.summary.total_occurrences, 6)

        # --- Stand-in for the engine: canned results ---
        canned = [
            _result("https://ok-site.com/page", Classification.OK,
                    http_status=200),
            _result("https://broken-site.com/missing", Classification.BROKEN,
                    http_status=404, error_detail="HTTP 404"),
            _result("https://throttled-site.com/api",
                    Classification.POSSIBLY_BLOCKED,
                    http_status=429, error_detail="HTTP 429",
                    likely_reason="Rate limited / 429"),
            _result("https://en.wikipedia.org/wiki/Python_(programming_language)",
                    Classification.OK, http_status=200),
            _result("https://dead-domain.example/x", Classification.BROKEN,
                    http_status=None,
                    error_detail="DNS error: no such host", final_url=None),
        ]

        # --- Phase 2: enrich (map results back to occurrences) ---
        enriched = enrich_results_with_occurrences(
            canned, extraction.occurrences_map
        )
        self.assertEqual(len(enriched), 5)

        # --- Phase 3: build rows and write ---
        broken_rows, blocked_rows = build_output_rows(enriched)
        self.assertEqual(len(broken_rows), 2)   # 2 broken URLs
        self.assertEqual(len(blocked_rows), 1)  # 1 blocked URL
        write_excel_report(broken_rows, blocked_rows, self.output_path)

        # --- Verify the output file ---
        out_wb = load_workbook(self.output_path)
        self.assertEqual(out_wb.sheetnames, [SHEET_BROKEN, SHEET_BLOCKED])

        # Broken sheet: rows sorted by domain.
        broken_ws = out_wb[SHEET_BROKEN]
        broken_urls = [broken_ws.cell(row=r, column=1).value
                       for r in range(2, broken_ws.max_row + 1)]
        self.assertEqual(broken_urls, [
            "https://broken-site.com/missing",   # b < d
            "https://dead-domain.example/x",
        ])

        # The duplicate "https://broken-site.com/missing" should produce
        # a Cell Location(s) field listing both occurrences. Input rows:
        #   A2 ok-site, A3 broken-site (first), A4 throttled, A5 wiki,
        #   A6 dead, A7 N/A (skip), A8 empty (skip), A9 broken-site (dup).
        first_broken_locations = broken_ws.cell(row=2, column=2).value
        self.assertIn("Sheet1!A3", first_broken_locations)
        self.assertIn("Sheet1!A9", first_broken_locations)

        # Blocked sheet: 1 row with Likely Reason column populated.
        blocked_ws = out_wb[SHEET_BLOCKED]
        self.assertEqual(blocked_ws.max_row, 2)  # header + 1 data row
        # Likely Reason is the last column (column 8).
        self.assertEqual(blocked_ws.cell(row=2, column=8).value,
                         "Rate limited / 429")

    def test_wikipedia_paren_url_survives_pipeline(self):
        """Regression guard: closing-paren bug from priority item #1."""
        _build_input_workbook(self.input_path, [
            ["https://en.wikipedia.org/wiki/Foo_(bar)"],
            ["https://en.wikipedia.org/wiki/Python_(programming_language)"],
        ])

        wb = open_workbook(self.input_path)
        selection = resolve_column(wb, "Sheet1", "URL", header_row=1)
        extraction = extract_urls(wb, selection, header_row=1)

        # Both URLs should make it through extraction with parens intact.
        self.assertIn(
            "https://en.wikipedia.org/wiki/Foo_(bar)",
            extraction.unique_urls,
        )
        self.assertIn(
            "https://en.wikipedia.org/wiki/Python_(programming_language)",
            extraction.unique_urls,
        )

        # And if those URLs come back as broken, the parens must survive
        # in the output sheet too.
        canned = [
            _result("https://en.wikipedia.org/wiki/Foo_(bar)",
                    Classification.BROKEN, http_status=404,
                    error_detail="HTTP 404"),
            _result("https://en.wikipedia.org/wiki/Python_(programming_language)",
                    Classification.BROKEN, http_status=404,
                    error_detail="HTTP 404"),
        ]
        enriched = enrich_results_with_occurrences(
            canned, extraction.occurrences_map
        )
        broken_rows, blocked_rows = build_output_rows(enriched)
        write_excel_report(broken_rows, blocked_rows, self.output_path)

        out_wb = load_workbook(self.output_path)
        ws = out_wb[SHEET_BROKEN]
        urls = [ws.cell(row=r, column=1).value
                for r in range(2, ws.max_row + 1)]
        self.assertIn("https://en.wikipedia.org/wiki/Foo_(bar)", urls)
        self.assertIn(
            "https://en.wikipedia.org/wiki/Python_(programming_language)",
            urls,
        )

    def test_pipeline_with_all_ok_produces_all_clear(self):
        _build_input_workbook(self.input_path, [
            ["https://a.com"],
            ["https://b.com"],
        ])
        wb = open_workbook(self.input_path)
        selection = resolve_column(wb, "Sheet1", "URL", header_row=1)
        extraction = extract_urls(wb, selection, header_row=1)

        canned = [
            _result("https://a.com", Classification.OK, http_status=200),
            _result("https://b.com", Classification.OK, http_status=200),
        ]
        enriched = enrich_results_with_occurrences(
            canned, extraction.occurrences_map
        )
        broken_rows, blocked_rows = build_output_rows(enriched)
        write_excel_report(broken_rows, blocked_rows, self.output_path)

        out_wb = load_workbook(self.output_path)
        self.assertEqual(out_wb.sheetnames, [SHEET_ALL_CLEAR])

    def test_cell_locations_list_all_occurrences(self):
        """If a URL appears in 3 cells, all 3 should be listed in the report."""
        _build_input_workbook(self.input_path, [
            ["https://repeat.com/x"],   # A2
            ["https://other.com"],       # A3
            ["https://repeat.com/x"],   # A4
            ["https://repeat.com/x"],   # A5
        ])
        wb = open_workbook(self.input_path)
        selection = resolve_column(wb, "Sheet1", "URL", header_row=1)
        extraction = extract_urls(wb, selection, header_row=1)

        canned = [
            _result("https://repeat.com/x", Classification.BROKEN,
                    http_status=404, error_detail="HTTP 404"),
            _result("https://other.com", Classification.OK, http_status=200),
        ]
        enriched = enrich_results_with_occurrences(
            canned, extraction.occurrences_map
        )
        broken_rows, _ = build_output_rows(enriched)
        write_excel_report(broken_rows, [], self.output_path)

        out_wb = load_workbook(self.output_path)
        ws = out_wb[SHEET_BROKEN]
        locations = ws.cell(row=2, column=2).value
        for cell_ref in ("Sheet1!A2", "Sheet1!A4", "Sheet1!A5"):
            self.assertIn(cell_ref, locations,
                          f"missing {cell_ref} in {locations!r}")


class PipelineHandlesMultiUrlCells(unittest.TestCase):
    """Cells with multiple URLs (comma/semicolon-separated)."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self.tmp.name)
        self.input_path = self.tmp_path / "input.xlsx"
        self.output_path = self.tmp_path / "output.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_two_urls_in_one_cell_both_extracted(self):
        _build_input_workbook(self.input_path, [
            ["https://first.com, https://second.com"],
        ])
        wb = open_workbook(self.input_path)
        selection = resolve_column(wb, "Sheet1", "URL", header_row=1)
        extraction = extract_urls(wb, selection, header_row=1)
        self.assertEqual(extraction.summary.unique_urls, 2)
        self.assertIn("https://first.com", extraction.unique_urls)
        self.assertIn("https://second.com", extraction.unique_urls)


if __name__ == "__main__":
    unittest.main(verbosity=2)
