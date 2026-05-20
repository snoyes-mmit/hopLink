"""Behavioral tests for the typed exceptions in `urlcheck.excel_input`.

Covers:
  - Every raise site uses the right typed exception class.
  - Exception attributes are populated (so callers can format their own
    messages without parsing str(e)).
  - Backwards-compatible class hierarchy: existing `except ValueError`
    and `except FileNotFoundError` paths in legacy callers still work.
  - The worker's `_friendly_excel_error` dispatches correctly for each
    type and uses the attributes (not str(e) substring matching).

These tests would have caught the brittleness of the old string-matching
approach. They will continue to pass if `excel_input.py` rephrases any
error message wording, since they assert on attributes, not text.

Run with:
    python -m unittest tests.test_excel_input_errors -v
"""

from __future__ import annotations

import io
import sys
import tempfile
import types
import unittest
from pathlib import Path

from openpyxl import Workbook

from urlcheck.excel_input import (
    ColumnNotFound,
    EmptyWorkbook,
    ExcelFileNotFound,
    ExcelInputError,
    HeaderNotFound,
    InvalidColumnReference,
    InvalidExcelFile,
    SheetNotFound,
    open_workbook,
    resolve_column,
)


# ---------------------------------------------------------------------------
# Stub PySide6 so we can also test worker._friendly_excel_error here.
# (Mirrors the stub installed in test_worker_logging.py — both files run
# in the same process during a `discover` run, and the first one to import
# wins. The duplication is intentional: each file is self-contained.)
# ---------------------------------------------------------------------------

def _install_pyside6_stub() -> None:
    if "PySide6" in sys.modules:
        return

    pyside6 = types.ModuleType("PySide6")
    qtcore = types.ModuleType("PySide6.QtCore")

    class _QThreadStub:
        def __init__(self, *a, **kw):
            pass

    class _QObjectStub:
        def __init__(self, *a, **kw):
            pass

    class _SignalStub:
        def __init__(self, *a, **kw):
            pass

        def emit(self, *a, **kw):
            pass

        def connect(self, *a, **kw):
            pass

    def _slot_stub(*a, **kw):
        def decorator(fn):
            return fn
        if len(a) == 1 and callable(a[0]) and not kw:
            return a[0]
        return decorator

    qtcore.QThread = _QThreadStub
    qtcore.QObject = _QObjectStub
    qtcore.Signal = _SignalStub
    qtcore.Slot = _slot_stub
    pyside6.QtCore = qtcore
    sys.modules["PySide6"] = pyside6
    sys.modules["PySide6.QtCore"] = qtcore


_install_pyside6_stub()
from urlcheck.gui import worker as worker_mod  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(path: Path, sheet_names: list[str], header: str = "URL") -> None:
    wb = Workbook()
    # Workbook() starts with one default sheet; rename/remove to match.
    default = wb.active
    if sheet_names:
        default.title = sheet_names[0]
        default["A1"] = header
        for extra in sheet_names[1:]:
            new = wb.create_sheet(extra)
            new["A1"] = header
    else:
        # An empty workbook is hard to create with openpyxl — you can't
        # delete every sheet — so callers that need this scenario should
        # mock instead.
        pass
    wb.save(path)


# ---------------------------------------------------------------------------
# Exception hierarchy: subclassing for backwards compatibility
# ---------------------------------------------------------------------------

class ExceptionHierarchy(unittest.TestCase):
    """Classes must be wired so legacy `except ValueError` / `except
    FileNotFoundError` paths still catch everything they used to.
    """

    def test_excel_input_error_is_value_error(self):
        self.assertTrue(issubclass(ExcelInputError, ValueError))

    def test_excel_file_not_found_is_filenotfounderror(self):
        # Critical for the GUI / CLI: they have `except FileNotFoundError`
        # branches that pre-date this change.
        self.assertTrue(issubclass(ExcelFileNotFound, FileNotFoundError))

    def test_excel_file_not_found_is_also_excel_input_error(self):
        self.assertTrue(issubclass(ExcelFileNotFound, ExcelInputError))

    def test_invalid_excel_file_is_value_error(self):
        # Old callers used `except (ValueError, FileNotFoundError)` —
        # confirming this keeps them working.
        self.assertTrue(issubclass(InvalidExcelFile, ValueError))

    def test_header_and_invalid_ref_are_column_not_found(self):
        self.assertTrue(issubclass(HeaderNotFound, ColumnNotFound))
        self.assertTrue(issubclass(InvalidColumnReference, ColumnNotFound))

    def test_column_not_found_is_excel_input_error(self):
        self.assertTrue(issubclass(ColumnNotFound, ExcelInputError))

    def test_every_exception_subclasses_value_error(self):
        """Old `except ValueError` should still catch every category."""
        for cls in (ExcelFileNotFound, InvalidExcelFile, EmptyWorkbook,
                    SheetNotFound, HeaderNotFound, InvalidColumnReference):
            with self.subTest(cls=cls):
                self.assertTrue(
                    issubclass(cls, ValueError),
                    f"{cls.__name__} must subclass ValueError for "
                    "backwards-compat with legacy callers.",
                )


# ---------------------------------------------------------------------------
# open_workbook
# ---------------------------------------------------------------------------

class OpenWorkbookErrors(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self.tmp.name)

    def tearDown(self):
        self.tmp.cleanup()

    def test_missing_file_raises_excel_file_not_found(self):
        missing = self.tmp_path / "nope.xlsx"
        with self.assertRaises(ExcelFileNotFound) as ctx:
            open_workbook(missing)
        # The exception must carry the offending path as an attribute.
        self.assertEqual(ctx.exception.path, missing)

    def test_missing_file_still_caught_by_filenotfounderror(self):
        """Legacy callers catching `FileNotFoundError` must keep working."""
        missing = self.tmp_path / "nope.xlsx"
        with self.assertRaises(FileNotFoundError):
            open_workbook(missing)

    def test_missing_file_still_caught_by_value_error(self):
        """Legacy callers catching `ValueError` must keep working."""
        missing = self.tmp_path / "nope.xlsx"
        with self.assertRaises(ValueError):
            open_workbook(missing)

    def test_non_xlsx_raises_invalid_excel_file(self):
        bogus = self.tmp_path / "not_really.xlsx"
        bogus.write_text("this is not a real xlsx", encoding="utf-8")
        with self.assertRaises(InvalidExcelFile) as ctx:
            open_workbook(bogus)
        self.assertEqual(ctx.exception.path, bogus)
        # cause_message should be populated with openpyxl's underlying detail.
        self.assertTrue(ctx.exception.cause_message,
                        "cause_message should not be empty")

    def test_invalid_excel_file_caught_by_value_error(self):
        bogus = self.tmp_path / "not_really.xlsx"
        bogus.write_text("nonsense", encoding="utf-8")
        with self.assertRaises(ValueError):
            open_workbook(bogus)

    def test_valid_workbook_opens_without_error(self):
        good = self.tmp_path / "ok.xlsx"
        _make_workbook(good, ["Sheet1"])
        wb = open_workbook(good)
        self.assertIn("Sheet1", wb.sheetnames)


# ---------------------------------------------------------------------------
# resolve_column
# ---------------------------------------------------------------------------

class ResolveColumnErrors(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        path = Path(self.tmp.name) / "wb.xlsx"
        _make_workbook(path, ["Sheet1", "Other"], header="URL")
        self.wb = open_workbook(path)

    def tearDown(self):
        self.tmp.cleanup()

    # ----- SheetNotFound -----

    def test_unknown_sheet_raises_sheet_not_found(self):
        with self.assertRaises(SheetNotFound) as ctx:
            resolve_column(self.wb, "DoesNotExist", "URL", header_row=1)
        self.assertEqual(ctx.exception.sheet, "DoesNotExist")
        self.assertEqual(
            sorted(ctx.exception.available),
            ["Other", "Sheet1"],
        )

    def test_unknown_sheet_caught_by_value_error(self):
        with self.assertRaises(ValueError):
            resolve_column(self.wb, "DoesNotExist", "URL", header_row=1)

    # ----- HeaderNotFound -----

    def test_unknown_header_raises_header_not_found(self):
        # "NotAColumn" isn't a header in row 1 and contains digits/punct
        # — wait, no, it's pure letters. The function will try header
        # lookup first (since header_row=1), fail, then fall back to
        # interpreting it as a column letter "NotAColumn", which IS a
        # valid letter sequence. So we need a name that's NOT a valid
        # column letter to force the header path.
        with self.assertRaises(HeaderNotFound) as ctx:
            resolve_column(self.wb, "Sheet1", "Vendor Name", header_row=1)
        self.assertEqual(ctx.exception.header, "Vendor Name")
        self.assertEqual(ctx.exception.sheet, "Sheet1")
        self.assertEqual(ctx.exception.header_row, 1)

    def test_unknown_header_is_column_not_found(self):
        # Tests that the new ColumnNotFound parent class works as a
        # category catch — callers can write `except ColumnNotFound` to
        # handle both bad-header and bad-letter cases together.
        with self.assertRaises(ColumnNotFound):
            resolve_column(self.wb, "Sheet1", "Vendor Name", header_row=1)

    # ----- InvalidColumnReference -----

    def test_zero_index_raises_invalid_column_reference(self):
        with self.assertRaises(InvalidColumnReference) as ctx:
            resolve_column(self.wb, "Sheet1", 0, header_row=1)
        self.assertEqual(ctx.exception.column_ref, 0)

    def test_negative_index_raises_invalid_column_reference(self):
        with self.assertRaises(InvalidColumnReference) as ctx:
            resolve_column(self.wb, "Sheet1", -5, header_row=1)
        self.assertEqual(ctx.exception.column_ref, -5)

    def test_zero_string_index_raises_invalid_column_reference(self):
        with self.assertRaises(InvalidColumnReference) as ctx:
            resolve_column(self.wb, "Sheet1", "0", header_row=1)
        self.assertEqual(ctx.exception.column_ref, "0")

    def test_empty_string_raises_invalid_column_reference(self):
        with self.assertRaises(InvalidColumnReference) as ctx:
            resolve_column(self.wb, "Sheet1", "", header_row=1)
        self.assertEqual(ctx.exception.column_ref, "")

    def test_non_letter_string_without_header_row(self):
        # When header_row is None, anything that's not a digit/letter
        # has no resolution path.
        with self.assertRaises(InvalidColumnReference) as ctx:
            resolve_column(self.wb, "Sheet1", "Vendor Name", header_row=None)
        self.assertEqual(ctx.exception.column_ref, "Vendor Name")

    # ----- Successful resolutions still work -----

    def test_header_name_resolves(self):
        sel = resolve_column(self.wb, "Sheet1", "URL", header_row=1)
        self.assertEqual(sel.column, 1)
        self.assertEqual(sel.header, "URL")

    def test_column_letter_resolves(self):
        sel = resolve_column(self.wb, "Sheet1", "A", header_row=1)
        # "A" is also the header name "URL" in row 1, but only "URL"
        # matches that header — so "A" falls through to letter parsing.
        self.assertEqual(sel.column, 1)

    def test_integer_index_resolves(self):
        sel = resolve_column(self.wb, "Sheet1", 1, header_row=1)
        self.assertEqual(sel.column, 1)


# ---------------------------------------------------------------------------
# worker._friendly_excel_error dispatch
# ---------------------------------------------------------------------------

class FriendlyExcelErrorDispatch(unittest.TestCase):
    """The worker must format each typed exception via attributes.

    These tests catch the exact regression the old string-matching code
    had: if someone rephrases the underlying message, the dispatcher
    keeps working as long as the EXCEPTION CLASS is right.
    """

    def test_excel_file_not_found_message(self):
        e = ExcelFileNotFound(Path("/tmp/missing.xlsx"))
        msg = worker_mod.Worker._friendly_excel_error(e)
        self.assertIn("Could not find", msg)
        self.assertIn("/tmp/missing.xlsx", msg)

    def test_invalid_excel_file_message(self):
        e = InvalidExcelFile(Path("/tmp/bogus.xlsx"), cause_message="x")
        msg = worker_mod.Worker._friendly_excel_error(e)
        self.assertIn(".xlsx", msg)
        # The friendly message must guide the user toward a fix.
        # We don't pin the exact wording — just that there's actionable
        # guidance referring to Excel's conversion step.
        self.assertIn("Excel", msg)

    def test_sheet_not_found_lists_available(self):
        e = SheetNotFound("DoesNotExist", ["Sheet1", "Vendors", "References"])
        msg = worker_mod.Worker._friendly_excel_error(e)
        self.assertIn("DoesNotExist", msg)
        # All three available sheets must be in the message.
        for s in ("Sheet1", "Vendors", "References"):
            self.assertIn(s, msg)

    def test_header_not_found_message(self):
        e = HeaderNotFound("Vendor Name", "Sheet1", 1)
        msg = worker_mod.Worker._friendly_excel_error(e)
        self.assertIn("Vendor Name", msg)
        self.assertIn("Sheet1", msg)

    def test_invalid_column_reference_message(self):
        e = InvalidColumnReference("@@", "Invalid column letter: '@@'")
        msg = worker_mod.Worker._friendly_excel_error(e)
        self.assertIn("@@", msg)

    def test_unknown_exception_falls_through_to_generic(self):
        # A plain RuntimeError should not crash the dispatcher.
        msg = worker_mod.Worker._friendly_excel_error(
            RuntimeError("something else went wrong")
        )
        self.assertIn("something else went wrong", msg)

    def test_empty_unknown_exception_uses_class_name(self):
        class WeirdError(Exception):
            pass

        msg = worker_mod.Worker._friendly_excel_error(WeirdError())
        self.assertIn("WeirdError", msg)

    def test_dispatch_resilient_to_message_changes(self):
        """The regression-guard test.

        Previously, `_friendly_excel_error` did `if "Sheet" in str(e) and
        "not in workbook" in str(e)`. Rewording the underlying message
        would silently break it. Now dispatch is by class, so we can
        prove resilience by constructing the exception with arbitrary
        message wording — the friendly output must still contain the
        relevant attribute data.
        """
        # Construct a SheetNotFound with deliberately weird wording (we
        # can't change the auto-generated message text without editing
        # excel_input, but we CAN verify the friendly output uses
        # `e.sheet` and `e.available` directly).
        e = SheetNotFound("MissingSheet", ["A", "B"])
        # Wipe the message — even if str(e) is empty, the friendly output
        # must still produce something useful from the attributes.
        e.args = ()  # str(e) is now "".
        msg = worker_mod.Worker._friendly_excel_error(e)
        self.assertIn("MissingSheet", msg,
                      "Dispatcher must read e.sheet, not parse str(e)")
        self.assertIn("A", msg)
        self.assertIn("B", msg)


# ---------------------------------------------------------------------------
# Backwards compat for existing callers
# ---------------------------------------------------------------------------

class LegacyCallerCompat(unittest.TestCase):
    """The original `cli.py` does `except (ValueError, FileNotFoundError)`.
    The original `worker.py` had `if isinstance(e, FileNotFoundError):`.
    Confirm those patterns still work after the change.
    """

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self.tmp.name)

    def tearDown(self):
        self.tmp.cleanup()

    def test_cli_style_combined_except(self):
        """Reproduces cli.py's `except (ValueError, FileNotFoundError)`."""
        # Missing file → ExcelFileNotFound → caught by FileNotFoundError.
        try:
            open_workbook(self.tmp_path / "x.xlsx")
        except (ValueError, FileNotFoundError):
            pass
        else:
            self.fail("Expected legacy except clause to catch the error")

        # Bogus file → InvalidExcelFile → caught by ValueError.
        bogus = self.tmp_path / "bogus.xlsx"
        bogus.write_text("not xlsx", encoding="utf-8")
        try:
            open_workbook(bogus)
        except (ValueError, FileNotFoundError):
            pass
        else:
            self.fail("Expected legacy except clause to catch the error")


if __name__ == "__main__":
    unittest.main(verbosity=2)
