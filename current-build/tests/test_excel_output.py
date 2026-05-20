"""Behavioral tests for `urlcheck.excel_output`.

These tests write actual .xlsx files to a tempdir and read them back with
openpyxl. They verify the *behavior* of the writer (sheet names, column
order, sort order, freeze panes, the All-Clear sheet, etc.) rather than
inspecting the source.

The `GoldenAgainstSampleReport` class compares the output of
`write_excel_report` against the bundled `tests/fixtures/sample_report.xlsx`
fixture, cell-for-cell. If anyone changes the writer's contract (column
order, sheet names, sort behavior), this will fail with a precise diff.

Run with:
    python -m unittest tests.test_excel_output -v
"""

from __future__ import annotations

import datetime as _dt
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from urlcheck.excel_output import (
    BLOCKED_COLUMNS,
    BROKEN_COLUMNS,
    SHEET_ALL_CLEAR,
    SHEET_BLOCKED,
    SHEET_BROKEN,
    auto_adjust_columns,
    build_output_rows,
    default_output_filename,
    sort_rows,
    write_excel_report,
)
from urlcheck.models import Classification, UrlCheckResult


FIXTURES = Path(__file__).resolve().parent / "fixtures"
SAMPLE_REPORT = FIXTURES / "sample_report.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_result(
    url,
    classification,
    *,
    http_status=None,
    error_detail=None,
    final_url="__USE_URL__",
    response_time_ms=None,
    checked_at_utc="",
    likely_reason=None,
    domain="",
):
    """Build a `UrlCheckResult` with sensible defaults for testing.

    `final_url` defaults to the same as `url` (the common case after a
    successful HEAD/GET). Pass `final_url=None` explicitly to represent a
    URL that never connected (DNS failure, etc.) — that produces an empty
    Final URL cell in the report.
    """
    if final_url == "__USE_URL__":
        final_url = url
    return UrlCheckResult(
        original_url=url,
        domain=domain or url.split("//", 1)[-1].split("/", 1)[0],
        classification=classification,
        http_status=http_status,
        error_detail=error_detail,
        final_url=final_url,
        response_time_ms=response_time_ms,
        likely_reason=likely_reason,
        checked_at_utc=checked_at_utc,
    )


# ---------------------------------------------------------------------------
# Column ordering and sheet names (spec)
# ---------------------------------------------------------------------------

class ColumnOrderingSpec(unittest.TestCase):

    def test_broken_columns_exact_order(self):
        self.assertEqual(
            BROKEN_COLUMNS,
            (
                "URL",
                "Cell Location(s)",
                "HTTP Code",
                "Error Detail",
                "Final URL",
                "Response Time (ms)",
                "Checked At (UTC)",
            ),
        )

    def test_blocked_columns_extend_broken_with_likely_reason(self):
        self.assertEqual(BLOCKED_COLUMNS, BROKEN_COLUMNS + ("Likely Reason",))

    def test_sheet_names_exact(self):
        self.assertEqual(SHEET_BROKEN, "Broken URLs")
        self.assertEqual(SHEET_BLOCKED, "Possibly Blocked")
        self.assertEqual(SHEET_ALL_CLEAR, "All Clear")


# ---------------------------------------------------------------------------
# Filename helper
# ---------------------------------------------------------------------------

class DefaultFilename(unittest.TestCase):

    def test_format_with_explicit_date(self):
        self.assertEqual(
            default_output_filename(_dt.date(2026, 5, 18)),
            "url_issues_2026-05-18.xlsx",
        )

    def test_format_today_shape(self):
        # Without arg: today's date in the right shape.
        name = default_output_filename()
        self.assertRegex(name, r"^url_issues_\d{4}-\d{2}-\d{2}\.xlsx$")


# ---------------------------------------------------------------------------
# build_output_rows: classification filtering
# ---------------------------------------------------------------------------

class BuildOutputRowsFilter(unittest.TestCase):

    def test_ok_results_dropped(self):
        items = [
            (_make_result("https://ok.com", Classification.OK,
                          http_status=200), ["Sheet1!A1"]),
            (_make_result("https://bad.com", Classification.BROKEN,
                          http_status=404), ["Sheet1!A2"]),
        ]
        broken, blocked = build_output_rows(items)
        self.assertEqual(len(broken), 1)
        self.assertEqual(len(blocked), 0)
        self.assertEqual(broken[0]["url"], "https://bad.com")

    def test_broken_goes_to_broken_blocked_to_blocked(self):
        items = [
            (_make_result("https://a.com", Classification.BROKEN,
                          http_status=500), ["Sheet1!A1"]),
            (_make_result("https://b.com", Classification.POSSIBLY_BLOCKED,
                          http_status=403, likely_reason="Cloudflare"),
             ["Sheet1!B1"]),
        ]
        broken, blocked = build_output_rows(items)
        self.assertEqual(len(broken), 1)
        self.assertEqual(len(blocked), 1)
        self.assertEqual(broken[0]["url"], "https://a.com")
        self.assertEqual(blocked[0]["url"], "https://b.com")
        # likely_reason only on the blocked row.
        self.assertNotIn("likely_reason", broken[0])
        self.assertEqual(blocked[0]["likely_reason"], "Cloudflare")

    def test_multiple_cell_locations_joined_with_comma_space(self):
        items = [
            (_make_result("https://a.com", Classification.BROKEN,
                          http_status=404),
             ["Sheet1!B12", "Sheet1!B47"]),
        ]
        broken, _ = build_output_rows(items)
        self.assertEqual(broken[0]["cell_locations"], "Sheet1!B12, Sheet1!B47")

    def test_accepts_enriched_result_objects(self):
        """Items can also be objects with .result and .locations attributes."""
        class FakeEnriched:
            def __init__(self, result, locations):
                self.result = result
                self.locations = locations

        items = [
            FakeEnriched(
                _make_result("https://a.com", Classification.BROKEN,
                             http_status=404),
                ["Sheet1!A1"],
            ),
        ]
        broken, _ = build_output_rows(items)
        self.assertEqual(broken[0]["url"], "https://a.com")

    def test_invalid_item_raises_typeerror(self):
        with self.assertRaises(TypeError):
            build_output_rows([12345])  # type: ignore[list-item]


# ---------------------------------------------------------------------------
# sort_rows: by (domain, url)
# ---------------------------------------------------------------------------

class SortRowsByDomain(unittest.TestCase):

    def _row(self, url):
        return {
            "url": url,
            "_domain": url.split("//", 1)[1].split("/", 1)[0].lower(),
            "_url_lower": url.lower(),
        }

    def test_sorts_by_domain_then_url(self):
        rows = [
            self._row("https://zebra.com/page1"),
            self._row("https://apple.com/page2"),
            self._row("https://apple.com/page1"),
            self._row("https://mango.com/page1"),
        ]
        sorted_rows = sort_rows(rows)
        self.assertEqual(
            [r["url"] for r in sorted_rows],
            [
                "https://apple.com/page1",
                "https://apple.com/page2",
                "https://mango.com/page1",
                "https://zebra.com/page1",
            ],
        )

    def test_case_insensitive_domain_grouping(self):
        rows = [
            self._row("https://APPLE.com/x"),
            self._row("https://apple.com/y"),
        ]
        # We pre-set _domain lowercased in the helper, so equal domains
        # cluster together regardless of original casing.
        sorted_rows = sort_rows(rows)
        domains = [r["_domain"] for r in sorted_rows]
        self.assertEqual(domains, ["apple.com", "apple.com"])


# ---------------------------------------------------------------------------
# write_excel_report: file structure (round-trip)
# ---------------------------------------------------------------------------

class WriteReportTwoSheets(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.out = Path(self.tmp.name) / "out.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_writes_both_sheets_with_correct_names(self):
        broken = [{
            "url": "https://a.com",
            "cell_locations": "Sheet1!A1",
            "http_status": 404, "error_detail": "HTTP 404",
            "final_url": "", "response_time_ms": "", "checked_at_utc": "",
            "_domain": "a.com", "_url_lower": "https://a.com",
        }]
        blocked = [{
            "url": "https://b.com",
            "cell_locations": "Sheet1!B1",
            "http_status": 403, "error_detail": "HTTP 403",
            "final_url": "", "response_time_ms": "", "checked_at_utc": "",
            "likely_reason": "Cloudflare",
            "_domain": "b.com", "_url_lower": "https://b.com",
        }]
        write_excel_report(broken, blocked, self.out)
        wb = load_workbook(self.out)
        self.assertEqual(wb.sheetnames, [SHEET_BROKEN, SHEET_BLOCKED])

    def test_header_row_matches_column_spec(self):
        broken = [{
            "url": "https://x.com", "cell_locations": "S!A1",
            "http_status": 404, "error_detail": "", "final_url": "",
            "response_time_ms": "", "checked_at_utc": "",
            "_domain": "x.com", "_url_lower": "https://x.com",
        }]
        write_excel_report(broken, [], self.out)
        wb = load_workbook(self.out)
        ws = wb[SHEET_BROKEN]
        headers = tuple(cell.value for cell in ws[1])
        self.assertEqual(headers, BROKEN_COLUMNS)

    def test_blocked_sheet_has_likely_reason_column(self):
        blocked = [{
            "url": "https://y.com", "cell_locations": "S!B1",
            "http_status": 403, "error_detail": "", "final_url": "",
            "response_time_ms": "", "checked_at_utc": "",
            "likely_reason": "Cloudflare bot protection",
            "_domain": "y.com", "_url_lower": "https://y.com",
        }]
        write_excel_report([], blocked, self.out)
        wb = load_workbook(self.out)
        # All-Clear sheet branch only fires if BOTH lists are empty, so
        # we must still get the two-sheet layout here.
        self.assertIn(SHEET_BLOCKED, wb.sheetnames)
        ws = wb[SHEET_BLOCKED]
        headers = tuple(cell.value for cell in ws[1])
        self.assertEqual(headers, BLOCKED_COLUMNS)
        # Final body cell should contain the likely-reason text.
        last_col = len(BLOCKED_COLUMNS)
        self.assertEqual(ws.cell(row=2, column=last_col).value,
                         "Cloudflare bot protection")

    def test_header_row_is_frozen(self):
        broken = [{
            "url": "https://x.com", "cell_locations": "S!A1",
            "http_status": 404, "error_detail": "", "final_url": "",
            "response_time_ms": "", "checked_at_utc": "",
            "_domain": "x.com", "_url_lower": "https://x.com",
        }]
        write_excel_report(broken, [], self.out)
        wb = load_workbook(self.out)
        ws = wb[SHEET_BROKEN]
        self.assertEqual(ws.freeze_panes, "A2")

    def test_returns_resolved_path(self):
        result = write_excel_report([], [], self.out)
        self.assertEqual(result, self.out)
        self.assertTrue(self.out.exists())

    def test_creates_parent_directory(self):
        nested = Path(self.tmp.name) / "nested" / "deeper" / "report.xlsx"
        write_excel_report([], [], nested)
        self.assertTrue(nested.exists())


class WriteReportAllClear(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.out = Path(self.tmp.name) / "out.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_empty_inputs_produce_all_clear_sheet(self):
        write_excel_report([], [], self.out)
        wb = load_workbook(self.out)
        self.assertEqual(wb.sheetnames, [SHEET_ALL_CLEAR])
        ws = wb[SHEET_ALL_CLEAR]
        cell_text = ws["A1"].value or ""
        self.assertIn("No URL issues found", cell_text)

    def test_all_clear_does_not_write_two_sheet_layout(self):
        write_excel_report([], [], self.out)
        wb = load_workbook(self.out)
        self.assertNotIn(SHEET_BROKEN, wb.sheetnames)
        self.assertNotIn(SHEET_BLOCKED, wb.sheetnames)


# ---------------------------------------------------------------------------
# Sort order in the written file
# ---------------------------------------------------------------------------

class WrittenFileSortOrder(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.out = Path(self.tmp.name) / "out.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_broken_rows_sorted_by_domain(self):
        # Build via build_output_rows so the sort keys are populated.
        items = [
            (_make_result("https://zebra.com/x", Classification.BROKEN,
                          http_status=404), ["S!A1"]),
            (_make_result("https://apple.com/y", Classification.BROKEN,
                          http_status=500), ["S!A2"]),
            (_make_result("https://mango.com/z", Classification.BROKEN,
                          http_status=404), ["S!A3"]),
        ]
        broken, _ = build_output_rows(items)
        write_excel_report(broken, [], self.out)
        wb = load_workbook(self.out)
        ws = wb[SHEET_BROKEN]
        urls_in_order = [ws.cell(row=r, column=1).value for r in (2, 3, 4)]
        self.assertEqual(urls_in_order, [
            "https://apple.com/y",
            "https://mango.com/z",
            "https://zebra.com/x",
        ])


# ---------------------------------------------------------------------------
# Golden comparison against the bundled sample_report.xlsx fixture
# ---------------------------------------------------------------------------

@unittest.skipUnless(SAMPLE_REPORT.exists(),
                     f"sample fixture not present at {SAMPLE_REPORT}")
class GoldenAgainstSampleReport(unittest.TestCase):
    """Reproduce the contents of `sample_report.xlsx` from scratch.

    The fixture contains 3 broken URLs and 3 possibly-blocked URLs with
    specific cell locations, status codes, error detail strings, final
    URLs, response times, and timestamps. Building the same inputs and
    writing them should produce a workbook identical in structure and
    data to the fixture (modulo formatting, which we don't compare).
    """

    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.TemporaryDirectory()
        cls.out = Path(cls.tmp.name) / "regenerated.xlsx"

        broken_items = [
            (_make_result(
                "https://api.olddata.com/v1/items",
                Classification.BROKEN,
                http_status=500,
                error_detail="HTTP 500",
                final_url="https://api.olddata.com/v1/items",
                response_time_ms=2540,
                checked_at_utc="2026-05-06T14:23:14Z",
            ), ["Sheet1!B89"]),
            (_make_result(
                "https://api.olddata.com/v1/users",
                Classification.BROKEN,
                http_status=404,
                error_detail="HTTP 404",
                final_url="https://api.olddata.com/v1/users",
                response_time_ms=145,
                checked_at_utc="2026-05-06T14:23:11Z",
            ), ["Sheet1!B12", "Sheet1!B47"]),
            (_make_result(
                "https://nonexistent-zzz.example/page",
                Classification.BROKEN,
                http_status=None,
                error_detail="DNS error: nodename nor servname provided",
                final_url=None,
                response_time_ms=512,
                checked_at_utc="2026-05-06T14:23:18Z",
            ), ["Vendors!C7"]),
        ]

        blocked_items = [
            (_make_result(
                "https://bigsite.com/article/1",
                Classification.POSSIBLY_BLOCKED,
                http_status=403,
                error_detail="HTTP 403 — blocked indicator in headers",
                final_url="https://bigsite.com/article/1",
                response_time_ms=87,
                checked_at_utc="2026-05-06T14:23:25Z",
                likely_reason="Cloudflare bot protection",
            ), ["Sheet1!B102", "References!D15"]),
            (_make_result(
                "https://bigsite.com/article/2",
                Classification.POSSIBLY_BLOCKED,
                http_status=429,
                error_detail="HTTP 429",
                final_url="https://bigsite.com/article/2",
                response_time_ms=92,
                checked_at_utc="2026-05-06T14:23:27Z",
                likely_reason="Rate limited / 429",
            ), ["Sheet1!B103"]),
            (_make_result(
                "https://retailer.example/product/abc",
                Classification.POSSIBLY_BLOCKED,
                http_status=403,
                error_detail="HTTP 403 — blocked indicator in body",
                final_url="https://retailer.example/product/abc",
                response_time_ms=156,
                checked_at_utc="2026-05-06T14:23:30Z",
                likely_reason="PerimeterX challenge",
            ), ["Sheet1!B201"]),
        ]

        broken_rows, blocked_rows = build_output_rows(
            broken_items + blocked_items
        )
        write_excel_report(broken_rows, blocked_rows, cls.out)

        cls.expected_wb = load_workbook(SAMPLE_REPORT)
        cls.actual_wb = load_workbook(cls.out)

    @classmethod
    def tearDownClass(cls):
        cls.tmp.cleanup()

    def test_sheet_names_match_sample(self):
        self.assertEqual(self.actual_wb.sheetnames, self.expected_wb.sheetnames)

    def test_broken_sheet_cells_match_sample(self):
        self._assert_sheets_equal(SHEET_BROKEN)

    def test_blocked_sheet_cells_match_sample(self):
        self._assert_sheets_equal(SHEET_BLOCKED)

    def _assert_sheets_equal(self, sheet_name: str):
        expected = self.expected_wb[sheet_name]
        actual = self.actual_wb[sheet_name]
        self.assertEqual(
            actual.max_row, expected.max_row,
            f"{sheet_name}: row count differs",
        )
        self.assertEqual(
            actual.max_column, expected.max_column,
            f"{sheet_name}: column count differs",
        )
        for row in range(1, expected.max_row + 1):
            for col in range(1, expected.max_column + 1):
                exp = expected.cell(row=row, column=col).value
                got = actual.cell(row=row, column=col).value
                # Normalize Nones to "" for empty cells — the writer stores
                # missing optional fields as "" and openpyxl reads truly
                # empty cells as None. Either is acceptable as long as both
                # are "no data".
                if exp in (None, ""):
                    self.assertIn(got, (None, ""),
                                  f"{sheet_name}!{expected.cell(row=row, column=col).coordinate}: "
                                  f"expected empty, got {got!r}")
                else:
                    self.assertEqual(
                        got, exp,
                        f"{sheet_name}!{expected.cell(row=row, column=col).coordinate}: "
                        f"expected {exp!r}, got {got!r}",
                    )


# ---------------------------------------------------------------------------
# auto_adjust_columns: column widths bounded by min/max
# ---------------------------------------------------------------------------

class AutoAdjustColumnsBounds(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.out = Path(self.tmp.name) / "out.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_columns_have_sensible_widths(self):
        # Build a real row with a very long URL — width should be capped.
        long_url = "https://example.com/" + ("a" * 500)
        items = [
            (_make_result(long_url, Classification.BROKEN, http_status=404),
             ["Sheet1!A1"]),
        ]
        broken, _ = build_output_rows(items)
        write_excel_report(broken, [], self.out)
        wb = load_workbook(self.out)
        ws = wb[SHEET_BROKEN]
        # First column = URL. Width should be the cap, not 500+.
        width = ws.column_dimensions["A"].width
        self.assertIsNotNone(width)
        self.assertLessEqual(width, 100,
                             "URL column width should be capped, not unbounded")
        self.assertGreaterEqual(width, 10,
                                "Width should be at least the floor (10)")


if __name__ == "__main__":
    unittest.main(verbosity=2)
